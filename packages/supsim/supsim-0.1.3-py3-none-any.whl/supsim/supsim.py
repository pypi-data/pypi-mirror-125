#!/usr/bin/env python

import sys
import os
import random
import math
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from sklearn.linear_model import LinearRegression
from argparse import ArgumentParser, RawDescriptionHelpFormatter
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.cell import column_index_from_string
from tqdm import trange
from .rand_poly import Poly, RandPoly
from .util import *

def validate_args(args):
    msg = []
    if args.count<1:
        msg.append('invalid value for count')
    if args.size<1:
        msg.append('invalid value for count')
    if args.abs_ry1<0 or args.abs_ry1>1:
        msg.append('invalid value for abs-ry1')
    if args.abs_ry2<0 or args.abs_ry2>1:
        msg.append('invalid value for abs-ry2')
    try:
        a,b = args.abs_r12.split(',')
        args.abs_r12a, args.abs_r12b = float(a), float(b)
    except ValueError:
        msg.append('invalid range for abs-r12')
    if args.abs_r12a<0 or args.abs_r12a>1 or args.abs_r12b<0 or args.abs_r12b>1 or args.abs_r12a>args.abs_r12b:
        msg.append('invalid range for abs-r12')
    if args.noise<0:
        msg.append('invalid value for noise')
    if args.outprob is not None:
        if args.outfile is None:
            msg.append('outprob requires outfile')
        else:
            if args.outprob<0 or args.outprob>1:
                msg.append('invalid value for outprob')
    if args.plots is None:
        args.plots=set()
    else:
        if args.plotfile is None:
            msg.append('plots requires plotfile')
        else:
            args.plots = set(args.plots.split(','))
            if not args.plots.issubset(set(['R2', 'B1', 'B2'])):
                msg.append('invalid choices for plots')
    #if args.dpi is None:
    #    args.dpi = 200
    #else:
    #    if args.dpi<1:
    #        msg.append('invalid dpi value')
    #if args.rtm_plots is None:
    #    args.rtm_plots=set()
    #else:
    #    if args.plotfile is None:
    #        msg.append('rtm-plots requires plotfile')
    #    else:
    #        try:
    #            args.rtm_plots = {1+(column_index_from_string(col)-1)//5: col for col in args.rtm_plots.split(',')}
    #        except ValueError as e:
    #            msg.append(str(e))
    return msg

example_text = '''Example of a correct command:
supsim --count 100 --size 50 --mean 50 --sd 5 --abs-ry1 1.0 --abs-ry2 1.0 --abs-r12 0.1,0.7 --noise 0.05 --outprob 0.5 --outfile data.xlsx --plots R2 --plotfile plots.png
'''

def main():
    # define parameters
    parser = ArgumentParser(prog='supsim', formatter_class=RawDescriptionHelpFormatter, description='Simulate two-predictor suppression', epilog=example_text)
    parser.add_argument('--enhancement', help='Leaving "enhancement" not specified, the command prints RTM\'s falling within all regions with or without suppression. While, specifying enhancement value (a proportion between 0 and 1), the command prints only those RTM\'s falling within enhancement regions which show at least "the specified value" or greater values of enhancement', metavar='PROPORTION', type=float)
    parser.add_argument('--count', help='Specifying an "integer", determines the "number of RTM\'s" to be simulated', metavar='INTEGER', type=int, required=True)
    parser.add_argument('--size', help='Specifying an "integer", determines the "sample size" for "x1", "x2", and "y" vectors', metavar='INTEGER', type=int, required=True)
    parser.add_argument('--mean', help='Specifying a "value", determines the "means" of the normal distributions of the two predictors: "x1" and "x2"', metavar='VALUE', type=float, required=True)
    parser.add_argument('--sd', help='Specifying a "value", determines the "SD\'s" of the normal distributions of the two predictors: "x1" and "x2"', metavar='VALUE', type=float, required=True)
    parser.add_argument('--abs-ry1', help='Specifying a "value" (between 0 and 1), determines the maximum absolute value allowed for ry1', metavar='VALUE', type=float, required=True)
    parser.add_argument('--abs-ry2', help='Specifying a "value" (between 0 and 1), determines the maximum absolute value allowed for ry2', metavar='VALUE', type=float, required=True)
    parser.add_argument('--abs-r12', help='Specifying a "range", determines the minimum and the maximum absolute values allowed for r12 (both between 0 and 1). Note that the "r12" or the "collinearity" range should be specified by two comma-separated values as follows: lowervalue,uppervalue (e.g., 0.1,0.8)', metavar='RANGE', required=True)
    parser.add_argument('--noise', help='Specifying a "coefficient", determines a coefficient to be multiplied by both the "mean" and the "SD" of the original y (yo\'s) to determine both the mean and the SD of the noise distribution to be used in generating "y" vector', metavar='COEFFICIENT', type=float, required=True)
    parser.add_argument('--seed', help='Specifying a "value", determines a "seed" for the random number generator which is needed for "reproducibility" or "replicability"', metavar='VALUE', type=int, default=0)
    parser.add_argument('--outprob', help='Specifying a "proportion" (between 0 and 1), determines a specific proportion of the RTM population to be randomly selected as "RTM sample". Note that random sampling from the simulated RTM\'s population requires specifying the --outfile', metavar='PROPORTION', type=float, default=1.0)
    parser.add_argument('--outfile', help='Selecting a "file name" is required for detailed spreadsheet report (example of a correct file name: data.xlsx)', metavar='FILENAME')
    parser.add_argument('--plots', help='List of values to be plotted (B1 B2 R2 or any combinations of them like R1,B2) (requires --plotfile)', metavar='LIST')
    #parser.add_argument('--rtm-plots', help='List of RTM ID\'s to be plotted as 3d scatter plots. Their ID is their Excel column position in the output file (requires --plotfile)', metavar='LIST')
    parser.add_argument('--plotfile', help='Selecting a "file name" is required for a graphical report (example of a correct graphical file name: plots.png)', metavar='FILENAME')
    #parser.add_argument('--dpi', help='Selecting a dpi value for plot files', metavar='INTEGER', type=int, default=200)
    # validate parameters
    args = parser.parse_args()
    msgs = validate_args(args)
    if len(msgs)>0:
        s = '\n'.join(msgs)
        print(f'supsim: error:\n{s}', file=sys.stderr)
        sys.exit(1)
    if args.outfile is not None:
        wb = Workbook()
        wb.save(args.outfile)
    data = {'reg-0': [], 'reg-1': [], 'reg-2': [], 'reg-3': [], 'reg-4': [], 'rev-0': [], 'rev-1': [], 'rev-2': [], 'rev-3': [], 'rev-4': [], 'cls-0': [], 'cls-1': [], 'cls-4': []}
    j = 0
    #print(args)
    seed(args.seed)
    for jj in trange(args.count):
        while True:
            p = RandPoly(1, 2, -10, 10)
            x1 = np.random.normal(size=args.size, loc=args.mean, scale=args.sd)
            r12 = random.uniform(args.abs_r12a, args.abs_r12b)*random_sign()
            x2 = correlated_normal(x1, r12, args.mean, args.sd)
            x = np.array([x1, x2]).T
            yo = p.eval_all(x)
            me, se = yo.mean()*args.noise, yo.std()*args.noise
            e = np.random.normal(size=args.size, loc=me, scale=se)
            y = yo+e
            ry2, ry1 = r(x2, y), r(x1, y)
            if not (abs(ry1) <= args.abs_ry1 and abs(ry2) <= args.abs_ry2):
                continue
            if not (abs(ry1)>abs(ry2) and r12<=ry1*ry2+math.sqrt((1-ry1**2)*(1-ry2**2)) and r12>=ry1*ry2-math.sqrt((1-ry1**2)*(1-ry2**2))):
                continue
            reg = LinearRegression().fit(x, y)
            pr = Poly(np.array([[0, reg.intercept_], reg.coef_]))
            yh = reg.predict(x)
            B2, B1 = beta(x2, y, x1), beta(x1, y, x2)
            R2, R2n = r2b(y, yh), r2b(y, yo)
            if args.enhancement is not None and not(R2>ry1**2+ry2**2+args.enhancement):
                continue
            if abs(ry2/ry1)<0.01 and abs(ry2)>=0 and abs(ry1)>0.01: # classical
                if r12<0 and np.sign(ry1)==np.sign(B1) and np.sign(ry2)==np.sign(B2) and R2>ry1**2+ry2**2: # region 1
                    region = 'cls-1'
                elif r12>0 and np.sign(ry1)==np.sign(B1) and np.sign(ry2)!=np.sign(B2) and R2>ry1**2+ry2**2: # region 4
                    region = 'cls-4'
                else: # should not happen
                    region = 'cls-0'
            else: # regular graph
                if np.sign(ry1)==np.sign(ry2): # regular
                    if r12<0 and np.sign(ry1)==np.sign(B1) and np.sign(ry2)==np.sign(B2) and R2>ry1**2+ry2**2: # region 1
                        region = 'reg-1'
                    elif r12>0 and np.sign(ry1)==np.sign(B1) and np.sign(ry2)==np.sign(B2) and R2<=ry1**2+ry2**2: # region 2
                        region = 'reg-2'
                    elif r12>0 and np.sign(ry1)==np.sign(B1) and np.sign(ry2)!=np.sign(B2) and R2<=ry1**2+ry2**2: # region 3
                        region = 'reg-3'
                    elif r12>0 and np.sign(ry1)==np.sign(B1) and np.sign(ry2)!=np.sign(B2) and R2>ry1**2+ry2**2: # region 4
                        region = 'reg-4'
                    else: # should not happen
                        region = 'reg-0'
                else: # reverse graph
                    if r12>0 and np.sign(ry1)==np.sign(B1) and np.sign(ry2)==np.sign(B2) and R2>ry1**2+ry2**2: # region 1
                        region = 'rev-1'
                    elif r12<0 and np.sign(ry1)==np.sign(B1) and np.sign(ry2)==np.sign(B2) and R2<=ry1**2+ry2**2: # region 2
                        region = 'rev-2'
                    elif r12<0 and np.sign(ry1)==np.sign(B1) and np.sign(ry2)!=np.sign(B2) and R2<=ry1**2+ry2**2: # region 3
                        region = 'rev-3'
                    elif r12<0 and np.sign(ry1)==np.sign(B1) and np.sign(ry2)!=np.sign(B2) and R2>ry1**2+ry2**2: # region 4
                        region = 'rev-4'
                    else: # should not happen
                        region = 'rev-0'
            data[region].append([ry1**2+ry2**2, ry1, ry2, r12, R2, B1, B2])     
            break
        if args.outfile is not None and np.random.uniform(0.0, 1.0)<args.outprob:
            j += 1
            s = f'RTM {j}\n'
            s += f'e: N({me:.3f}, {se:.3f})\n'
            s += f'r12: {r12:.3f}\n'
            s += f'ry2: {ry2:.3f}\n'
            s += f'ry1: {ry1:.3f}\n'
            s += f'B2: {B2:.3f}\n'
            s += f'B1: {B1:.3f}\n'
            s += f'R2: {R2:.3f}\n'
            s += f'R2 (rand poly): {R2n:.3f}\n'
            s += f'region: {region}\n'
            #s += f'linear: {yesno[p.is_linear()]}\n'
            s += f'regression: {pr}\n'
            s += f'equation: {p}'
            b = r12
            a = R2
            c = ry1**2+ry2**2
            d = B1/ry1
            e = B2/ry2
            f = ry1
            g = ry2
            h = B1
            k = B2
            l = R2-(ry1**2+ry2**2)
            ws = wb.active
            ws.merge_cells(start_row=1, end_row=1, start_column=j*5-4, end_column=j*5)
            ws.row_dimensions[1].height = 188
            ws.cell(row=1, column=j*5-4).alignment = Alignment(horizontal='general', vertical='top', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)
            ws.cell(row=1, column=j*5-4, value=s)
            ws.cell(row=2, column=j*5-4, value='R2')
            ws.cell(row=2, column=j*5-3, value='r12')
            ws.cell(row=2, column=j*5-2, value='ry1**2+ry2**2')
            ws.cell(row=2, column=j*5-1, value='B1/ry1')
            ws.cell(row=2, column=j*5, value='B2/ry2')
            ws.cell(row=3, column=j*5-4, value=a)
            ws.cell(row=3, column=j*5-3, value=b)
            ws.cell(row=3, column=j*5-2, value=c)
            ws.cell(row=3, column=j*5-1, value=d)
            ws.cell(row=3, column=j*5, value=e)
            ws.cell(row=4, column=j*5-4, value='ry1')
            ws.cell(row=4, column=j*5-3, value='ry2')
            ws.cell(row=4, column=j*5-2, value='B1')
            ws.cell(row=4, column=j*5-1, value='B2')
            ws.cell(row=4, column=j*5, value='R2-(ry1**2+ry2**2)')
            ws.cell(row=5, column=j*5-4, value=f)
            ws.cell(row=5, column=j*5-3, value=g)
            ws.cell(row=5, column=j*5-2, value=h)
            ws.cell(row=5, column=j*5-1, value=k)
            ws.cell(row=5, column=j*5, value=l)
            ws.cell(row=6, column=j*5-4, value='x1')
            ws.cell(row=6, column=j*5-3, value='x2')
            ws.cell(row=6, column=j*5-2, value='y')
            ws.cell(row=6, column=j*5-1, value='yo')
            ws.cell(row=6, column=j*5, value='yh') 
            for i in range(len(y)):
                ws.cell(row=i+7, column=j*5-4, value=x1[i])
                ws.cell(row=i+7, column=j*5-3, value=x2[i])
                ws.cell(row=i+7, column=j*5-2, value=y[i])
                ws.cell(row=i+7, column=j*5-1, value=yo[i])
                ws.cell(row=i+7, column=j*5, value=yh[i])
            ws.freeze_panes = ws['A3']
            wb.save(args.outfile)
            #if args.plotfile is not None and j in args.rtm_plots:
            #    fig = plt.figure()
            #    xx1, xx2 = np.meshgrid(np.linspace(min(x1), max(x1), 100), np.linspace(min(x2), max(x2), 100))
            #    xx = np.concatenate((xx1.reshape(-1, 1), xx2.reshape(-1, 1)), axis=1)
            #    yy = reg.predict(xx)
            #    ax = fig.add_subplot(111, projection='3d')
            #    ax.view_init(elev=10, azim=120)
            #    ax.set_zlim(2, 12)
            #    ax.plot_surface(xx2, xx1, yy.reshape(xx2.shape), alpha=0.2)
            #    #ax.plot_wireframe(xx2, xx1, yy.reshape(xx2.shape), color='red', alpha=0.2)
            #    ax.scatter(x2, x1, y, marker='o')
            #    ax.set_xlabel('x2')
            #    ax.set_ylabel('x1')
            #    ax.set_zlabel('y')
            #    fname = append_filename(args.plotfile, args.rtm_plots[j])
            #    plt.savefig(fname, bbox_inches='tight')
            #    print(f'Saved file: {os.path.abspath(fname)}')
    regions = {k: pd.DataFrame(data=v, columns=['sum_r12^2', 'ry1', 'ry2', 'r12', 'R2', 'B1', 'B2']) for k, v in data.items()}
    counts = {k: len(v) for k, v in regions.items()}
    percents = {k: 100*v/sum(counts.values()) for k, v in counts.items()}
    if args.outfile is not None:
        wb.close()
        print(f'Saved file: {os.path.abspath(args.outfile)}')
    if args.plotfile is not None:
        # regular plots
        fig, (ax1, ax2, ax3, ax4) = plt.subplots(1, 4, sharey=True)
        fig.subplots_adjust(wspace=0)
        fig.set_figwidth(12)
        fig.set_figheight(6)
        ylab = ''
        if 'R2' in args.plots:
            ax1.scatter(x=regions['reg-1'].r12, y=regions['reg-1'].R2, label='$R^2$', s=2)
            ax2.scatter(x=regions['reg-2'].r12, y=regions['reg-2'].R2, label='$R^2$', s=2)
            ax3.scatter(x=regions['reg-3'].r12, y=regions['reg-3'].R2, label='$R^2$', s=2)
            ax4.scatter(x=regions['reg-4'].r12, y=regions['reg-4'].R2, label='$R^2$', s=2)
            ylab += '$R^2$\n'
        if 'B1' in args.plots:
            ax1.scatter(x=regions['reg-1'].r12, y=regions['reg-1'].B1, label='$\\beta_1$', s=2)
            ax2.scatter(x=regions['reg-2'].r12, y=regions['reg-2'].B1, label='$\\beta_1$', s=2)
            ax3.scatter(x=regions['reg-3'].r12, y=regions['reg-3'].B1, label='$\\beta_1$', s=2)
            ax4.scatter(x=regions['reg-4'].r12, y=regions['reg-4'].B1, label='$\\beta_1$', s=2)
            ylab += '$\\beta_1$\n'
        if 'B2' in args.plots:
            ax1.scatter(x=regions['reg-1'].r12, y=regions['reg-1'].B2, label='$\\beta_2$', s=2)
            ax2.scatter(x=regions['reg-2'].r12, y=regions['reg-2'].B2, label='$\\beta_2$', s=2)
            ax3.scatter(x=regions['reg-3'].r12, y=regions['reg-3'].B2, label='$\\beta_2$', s=2)
            ax4.scatter(x=regions['reg-4'].r12, y=regions['reg-4'].B2, label='$\\beta_2$', s=2)
            ylab += '$\\beta_2$\n'
        ax1.set_ylabel(ylab, rotation=0, labelpad=24, size=14)
        ax1.title.set_text('Region 1\nEnhancement (%.2f%%)' % percents['reg-1'])
        ax2.title.set_text('Region 2\nRedundancy (%.2f%%)' % percents['reg-2'])
        ax3.title.set_text('Region 3\nSuppression (%.2f%%)' % percents['reg-3'])
        ax4.title.set_text('Region 4\nEnhancement (%.2f%%)' % percents['reg-4'])
        ax2.annotate('$r_{12}$', xy=(1, 0), xytext=(-15, -45), ha='left', va='top', xycoords='axes fraction', textcoords='offset points', fontsize=12, annotation_clip=False)
        ax1.set_xlim(-1, 0)
        ax2.set_xlim(0, 1)
        ax3.set_xlim(0, 1)
        ax4.set_xlim(0, 1)
        ax2.tick_params(labelleft=False, left=False, pad=20)
        ax3.tick_params(labelleft=False, left=False)
        ax4.tick_params(labelleft=False, left=False, pad=20)
        handles, labels = ax4.get_legend_handles_labels()
        fig.legend(handles, labels, ncol=3, bbox_to_anchor=(0.53, 1.1))
        fname = append_filename(args.plotfile, 'regular')
        plt.savefig(fname, bbox_inches='tight')
        print(f'Saved file: {os.path.abspath(fname)}')
        # reverse plots
        fig, (ax4, ax3, ax2, ax1) = plt.subplots(1, 4, sharey=True)
        fig.subplots_adjust(wspace=0)
        fig.set_figwidth(12)
        fig.set_figheight(6)
        ylab = ''
        if 'R2' in args.plots:
            ax1.scatter(x=regions['rev-1'].r12, y=regions['rev-1'].R2, label='$R^2$', s=2)
            ax2.scatter(x=regions['rev-2'].r12, y=regions['rev-2'].R2, label='$R^2$', s=2)
            ax3.scatter(x=regions['rev-3'].r12, y=regions['rev-3'].R2, label='$R^2$', s=2)
            ax4.scatter(x=regions['rev-4'].r12, y=regions['rev-4'].R2, label='$R^2$', s=2)
            ylab += '$R^2$\n'
        if 'B1' in args.plots:
            ax1.scatter(x=regions['rev-1'].r12, y=regions['rev-1'].B1, label='$\\beta_1$', s=2)
            ax2.scatter(x=regions['rev-2'].r12, y=regions['rev-2'].B1, label='$\\beta_1$', s=2)
            ax3.scatter(x=regions['rev-3'].r12, y=regions['rev-3'].B1, label='$\\beta_1$', s=2)
            ax4.scatter(x=regions['rev-4'].r12, y=regions['rev-4'].B1, label='$\\beta_1$', s=2)
            ylab += '$\\beta_1$\n'
        if 'B2' in args.plots:
            ax1.scatter(x=regions['rev-1'].r12, y=regions['rev-1'].B2, label='$\\beta_2$', s=2)
            ax2.scatter(x=regions['rev-2'].r12, y=regions['rev-2'].B2, label='$\\beta_2$', s=2)
            ax3.scatter(x=regions['rev-3'].r12, y=regions['rev-3'].B2, label='$\\beta_2$', s=2)
            ax4.scatter(x=regions['rev-4'].r12, y=regions['rev-4'].B2, label='$\\beta_2$', s=2)
            ylab += '$\\beta_2$\n'
        ax4.set_ylabel(ylab, rotation=0, labelpad=24, size=14)
        ax1.title.set_text('Region 1\nEnhancement (%.2f%%)' % percents['rev-1'])
        ax2.title.set_text('Region 2\nRedundancy (%.2f%%)' % percents['rev-2'])
        ax3.title.set_text('Region 3\nSuppression (%.2f%%)' % percents['rev-3'])
        ax4.title.set_text('Region 4\nEnhancement (%.2f%%)' % percents['rev-4'])
        ax2.annotate('$r_{12}$', xy=(0, 0), xytext=(-15, -45), ha='left', va='top', xycoords='axes fraction', textcoords='offset points', fontsize=12, annotation_clip=False)
        ax1.set_xlim(0, 1)
        ax2.set_xlim(-1, 0)
        ax3.set_xlim(-1, 0)
        ax4.set_xlim(-1, 0)
        ax3.tick_params(labelleft=False, left=False, pad=20)
        ax2.tick_params(labelleft=False, left=False)
        ax1.tick_params(labelleft=False, left=False, pad=20)
        handles, labels = ax4.get_legend_handles_labels()
        fig.legend(handles, labels, ncol=3, bbox_to_anchor=(0.53, 1.1))
        fname = append_filename(args.plotfile, 'reverse')
        plt.savefig(fname, bbox_inches='tight')
        print(f'Saved file: {os.path.abspath(fname)}')
        # classical plots
        fig, (ax1, ax4) = plt.subplots(1, 2, sharey=True)
        fig.subplots_adjust(wspace=0)
        fig.set_figwidth(12)
        fig.set_figheight(6)
        ylab = ''
        if 'R2' in args.plots:
            ax1.scatter(x=regions['cls-1'].r12, y=regions['cls-1'].R2, label='$R^2$', s=2)
            ax4.scatter(x=regions['cls-4'].r12, y=regions['cls-4'].R2, label='$R^2$', s=2)
            ylab += '$R^2$\n'
        if 'B1' in args.plots:
            ax1.scatter(x=regions['cls-1'].r12, y=regions['cls-1'].B1, label='$\\beta_1$', s=2)
            ax4.scatter(x=regions['cls-4'].r12, y=regions['cls-4'].B1, label='$\\beta_1$', s=2)
            ylab += '$\\beta_1$\n'
        if 'B2' in args.plots:
            ax1.scatter(x=regions['cls-1'].r12, y=regions['cls-1'].B2, label='$\\beta_2$', s=2)
            ax4.scatter(x=regions['cls-4'].r12, y=regions['cls-4'].B2, label='$\\beta_2$', s=2)
            ylab += '$\\beta_2$\n'
        ax1.set_ylabel(ylab, rotation=0, labelpad=24, size=14)
        ax1.title.set_text('Region 1\nEnhancement (%.2f%%)' % percents['cls-1'])
        ax4.title.set_text('Region 4\nEnhancement (%.2f%%)' % percents['cls-4'])
        ax1.annotate('$r_{12}$', xy=(1, 0), xytext=(-15, -45), ha='left', va='top', xycoords='axes fraction', textcoords='offset points', fontsize=12, annotation_clip=False)
        ax1.set_xlim(-1, 0)
        ax4.set_xlim(0, 1)
        ax4.tick_params(labelleft=False, left=False, pad=20)
        handles, labels = ax4.get_legend_handles_labels()
        fig.legend(handles, labels, ncol=3, bbox_to_anchor=(0.53, 1.1))
        fname = append_filename(args.plotfile, 'classical')
        plt.savefig(fname, bbox_inches='tight')
        print(f'Saved file: {os.path.abspath(fname)}')

if __name__ == '__main__':
    sys.exit(main())
